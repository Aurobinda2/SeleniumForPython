import openpyxl
path="E:\Sample-Sales-Data.xlsx"
workBook=openpyxl.load_workbook(path)
sheet=workBook.active
max_row=sheet.max_row
max_coloum=sheet.max_column

#Read the data from Excel


for r in range(1,max_row+1):
    for c in range(1,max_coloum+1):
        print(sheet.cell(row=r,column=c).value,end=' ')
    print()

# write Data to Excel

path2="E:\Test12"

workBook1=openpyxl.load_workbook(path)
sheet1=workBook1.active
for r in range(1,5):
    for c in range(1,3):
        sheet1.cell(row=r,column=c).value='Aurobinda'

workBook1.save(path2)
