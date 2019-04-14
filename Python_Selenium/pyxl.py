import openpyxl

#Function return the row count
def row_count(path,sheet):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheet)
    return(sheet.max_row)

#Function gives the column count

def column_count(path,sheet):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheet)
    return(sheet.max_column)

#Function read the data from excel sheet
def read_data(path,sheet,row,col):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheet)
    return(sheet.cell(row=row,column=col).value)

#Function write the data to excel  sheet
def write_data(path,sheet,row,col,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheet)
    sheet.cell(row=row,column=col).value=data
    workbook.save(path)


